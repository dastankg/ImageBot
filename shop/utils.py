from django.core.checks import messages
from django.http import HttpResponse
from django.shortcuts import render
from django import forms
from openpyxl import Workbook
from datetime import datetime
from post.models import Post, PostAgent
from django.contrib import admin
from django.utils.timezone import localtime


class DateRangeRegionForm(forms.Form):
    start_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "start_date"}),
        required=False,
        label="С даты (День/Месяц/Год)",
    )
    end_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "end_date"}),
        required=False,
        label="По дату (День/Месяц/Год)",
    )
    region = forms.CharField(
        widget=forms.TextInput(attrs={"name": "region"}),
        required=False,
        label="Регион",
    )


class DateRangeForm(forms.Form):
    start_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "start_date"}),
        required=False,
        label="С даты (День/Месяц/Год)",
    )
    end_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "end_date"}),
        required=False,
        label="По дату (День/Месяц/Год)",
    )


def export_posts_to_excel(modeladmin, request, queryset):
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")
    region = request.POST.get("region")

    if "apply" not in request.POST:
        form = DateRangeRegionForm()
        context = {
            "queryset": queryset,
            "form": form,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, "export_date_range.html", context)

    if start_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
    if end_date:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="posts.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    columns = [
        "ID",
        "Магазин",
        "Изображение",
        "Адрес",
        "Адрес магазина",
        "Регион",
        "Дата",
        "Время",
    ]
    ws.append(columns)

    if region:
        queryset = queryset.filter(region=region)

    for shop in queryset:
        posts = Post.objects.filter(shop=shop)

        if start_date:
            posts = posts.filter(created__gte=start_date)
        if end_date:
            posts = posts.filter(created__lte=end_date)

        for post in posts:
            local_created = localtime(post.created)
            ws.append(
                [
                    post.id,
                    shop.shop_name,
                    f"http://139.59.2.151:8000/media/{post.image}",
                    post.address,
                    shop.address,
                    shop.region,
                    local_created.strftime("%Y-%m-%d"),
                    local_created.strftime("%H:%M:%S"),
                ]
            )

        if posts:
            ws.append([])

    wb.save(response)
    return response


def export_agent_posts_to_excel(modeladmin, request, queryset):
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")

    if not queryset:
        modeladmin.message_user(
            request, "Не выбрано ни одного агента", level=messages.ERROR
        )
        return
    if "apply" not in request.POST:
        form = DateRangeForm()
        context = {
            "queryset": queryset,
            "form": form,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, "export_date_range1.html", context)

    if start_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
    if end_date:
        end_date = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="agent_posts.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Posts"

    columns = [
        "№",
        "Дата",
        "ФИО",
        "Название магазина",
        "Гео.",
        "Приход",
        "Уход",
        "Продолжительность",
        "Фото ТМ до",
        "Фото ТМ после",
        "Кол-во ТМ",
        "Кол-во ДПМ",
        "Фото ДПМ",
    ]
    ws.append(columns)

    row_number = 1

    for agent in queryset:
        posts = PostAgent.objects.filter(agent=agent).order_by("created")

        if start_date:
            posts = posts.filter(created__gte=start_date)
        if end_date:
            posts = posts.filter(created__lte=end_date)

        posts_list = list(posts)

        if not posts_list:
            continue

        blocks = []
        current_block = []
        current_shop = None

        for post in posts_list:
            if current_shop != post.shop:
                if current_block:
                    blocks.append(current_block)
                current_block = [post]
                current_shop = post.shop
            else:
                current_block.append(post)

        if current_block:
            blocks.append(current_block)

        for block in blocks:
            if not block:
                continue

            first_post = block[0]
            last_post = block[-1]

            arrival_time = localtime(first_post.created)
            departure_time = localtime(last_post.created)

            duration_seconds = (last_post.created - first_post.created).total_seconds()
            duration_minutes = int(duration_seconds / 60)
            hours = duration_minutes // 60
            minutes = duration_minutes % 60
            duration_str = f"{hours:02d}:{minutes:02d}"

            tm_before_photos = [p for p in block if p.post_type == "ТМ до"]
            tm_after_photos = [p for p in block if p.post_type == "ТМ после"]
            dpm_photos = [p for p in block if p.post_type == "ДПМ"]

            tm_before_count = len(tm_before_photos)
            tm_after_count = len(tm_after_photos)
            dpm_count = len(dpm_photos)
            total_tm_count = tm_before_count + tm_after_count

            def get_photo_urls(photos):
                if not photos:
                    return ""
                urls = []
                for photo in photos:
                    if photo.image:
                        url = request.build_absolute_uri(photo.image.url)
                        urls.append(url)
                return "\n".join(urls) if urls else ""

            tm_before_urls = get_photo_urls(tm_before_photos)
            tm_after_urls = get_photo_urls(tm_after_photos)
            dpm_urls = get_photo_urls(dpm_photos)

            row_data = [
                row_number,
                arrival_time.strftime("%d.%m.%y"),
                agent.agent_name,
                first_post.shop,
                first_post.address or "",
                arrival_time.strftime("%H:%M"),
                departure_time.strftime("%H:%M"),
                duration_str,
                tm_before_urls if tm_before_count > 0 else "",
                tm_after_urls if tm_after_count > 0 else "",
                total_tm_count,
                dpm_count,
                dpm_urls if dpm_count > 0 else "",
            ]

            ws.append(row_data)
            row_number += 1

        if blocks:
            ws.append([])

        if blocks:
            day_start = localtime(posts_list[0].created)
            day_end = localtime(posts_list[-1].created)
            total_day_duration = (
                posts_list[-1].created - posts_list[0].created
            ).total_seconds()

            total_work_duration = sum(
                [
                    (block[-1].created - block[0].created).total_seconds()
                    for block in blocks
                ]
            )

            work_percentage = (
                (total_work_duration / total_day_duration * 100)
                if total_day_duration > 0
                else 0
            )

            total_tm_before = sum(
                len([p for p in block if p.post_type == "ТМ до"]) for block in blocks
            )
            total_tm_after = sum(
                len([p for p in block if p.post_type == "ТМ после"]) for block in blocks
            )
            total_dpm = sum(
                len([p for p in block if p.post_type == "ДПМ"]) for block in blocks
            )
            total_photos = total_tm_before + total_tm_after + total_dpm

            summary_row = [
                "",
                "Итого за день",
                f"Начало: {day_start.strftime('%H:%M')}",
                f"Конец: {day_end.strftime('%H:%M')}",
                f"Общее время: {int(total_day_duration / 3600):02d}:{int((total_day_duration % 3600) / 60):02d}",
                f"Рабочее время: {int(total_work_duration / 3600):02d}:{int((total_work_duration % 3600) / 60):02d}",
                f"Эффективность: {work_percentage:.1f}%",
                f"ТМ до: {total_tm_before}",
                f"ТМ после: {total_tm_after}",
                f"Всего ТМ: {total_tm_before + total_tm_after}",
                f"ДПМ: {total_dpm}",
                f"Всего фото: {total_photos}",
                "",
            ]
            ws.append(summary_row)
            ws.append([])

    column_widths = [5, 12, 20, 25, 30, 10, 10, 15, 50, 50, 12, 12, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    wb.save(response)
    return response
